#!/usr/bin/env python3
"""
Export firmware-built-in recipes from ESP32 reader to Excel.
Data extracted from: NFC_recipes.c, NFC_recipes.h, NFC_reader.h.
No firmware behavior changed; this script is read-only extraction.
"""
from pathlib import Path
import sys

# Ensure we can find openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# --- Data from NFC_recipes.c / NFC_recipes.h (ProcessTypes, Alcohol, NonAlcohol) ---
PROCESS_TYPES = [
    (0, "ToStorageGlass"),
    (1, "StorageAlcohol"),
    (2, "StorageNonAlcohol"),
    (3, "Shaker"),
    (4, "Cleaner"),
    (5, "SodaMake"),
    (6, "ToCustomer"),
    (7, "Transport"),
    (8, "Buffer"),
]

ALCOHOL_TYPE = [(0, "Vodka"), (1, "Rum"), (2, "Goralka")]
NON_ALCOHOL_TYPE = [(0, "Water"), (1, "Cola")]

# --- Full recipes from GetCardInfoByNumber (NFC_recipes.c) ---
# RecipeNumber -> (RecipeName, NumSteps, Description, (step_type, param) list, ActualBudget if set)
RECIPES = [
    {
        "RecipeNumber": 0,
        "RecipeName": "Vodka s vodou (20+80ml)",
        "NumSteps": 5,
        "Description": "Generated in GetCardInfoByNumber case 0. Comments in code: Cisteni 5s, Vodka 20ml, Drink k zakaznikovi, Cisteni 20s, Navrat do skladu.",
        "SourceLocation": "NFC_recipes.c:141-153 GetCardInfoByNumber",
        "Steps": [
            (4, 5),   # StorageNonAlcohol Water 5
            (1, 20),  # StorageAlcohol Vodka 20
            (4, 0),   # StorageNonAlcohol Water 0
            (5, 20),  # StorageNonAlcohol Cola 20
            (8, 0),   # ToStorageGlass
        ],
        "ActualBudget": None,
    },
    {
        "RecipeNumber": 1,
        "RecipeName": "Rum s colou (40+60ml)",
        "NumSteps": 6,
        "Description": "Generated in GetCardInfoByNumber case 1. ActualBudget=200.",
        "SourceLocation": "NFC_recipes.c:154-166 GetCardInfoByNumber",
        "Steps": [
            (7, 5),   # Cleaner 5s
            (2, 40),  # StorageAlcohol Rum 40
            (5, 60),  # StorageNonAlcohol Cola 60
            (9, 0),   # ToCustomer (case 9)
            (7, 20),  # Cleaner 20s
            (8, 0),   # ToStorageGlass
        ],
        "ActualBudget": 200,
    },
    {
        "RecipeNumber": 2,
        "RecipeName": "Vodka s vodou (20+80ml)",
        "NumSteps": 5,
        "Description": "Same steps as recipe 0; generated in GetCardInfoByNumber case 2.",
        "SourceLocation": "NFC_recipes.c:167-177 GetCardInfoByNumber",
        "Steps": [
            (4, 5),
            (1, 20),
            (4, 0),
            (5, 20),
            (8, 0),
        ],
        "ActualBudget": None,
    },
    {
        "RecipeNumber": 3,
        "RecipeName": "Návrat do skladu",
        "NumSteps": 1,
        "Description": "Return to storage; single step ToStorageGlass. Used in State_Mimo_NastaveniNaPresunDoSkladu (app.c:493).",
        "SourceLocation": "NFC_recipes.c:179-186 GetCardInfoByNumber",
        "Steps": [(8, 0)],
        "ActualBudget": None,
    },
    {
        "RecipeNumber": 255,
        "RecipeName": "Empty/Default",
        "NumSteps": 0,
        "Description": "GetCardInfoByNumber default case; generates empty card (no steps). Not a valid selectable recipe in app.",
        "SourceLocation": "NFC_recipes.c:187-191 GetCardInfoByNumber (default)",
        "Steps": [],
        "ActualBudget": None,
    },
]

# Step type number -> (TypeOfProcess enum value, ParameterProcess1 meaning, label)
# From GetRecipeStepByNumber switch: 1-10. ParameterProcess1 for 1-5 is alcohol/non-alcohol subtype; for 6,7 it's duration; for 10 it's ProcessCellID.
STEP_TYPE_INFO = {
    1: (1, "Vodka (AlcoholType)", "Vodka(1) ml"),
    2: (1, "Rum (AlcoholType)", "Rum(2) ml"),
    3: (1, "Goralka (AlcoholType)", "Goralka(3) ml"),
    4: (2, "Water (NonAlcoholType)", "Voda/Water(4) ml"),
    5: (2, "Cola (NonAlcoholType)", "Cola(5) ml"),
    6: (3, "duration s", "Protrepani/Shaker s"),
    7: (4, "duration s", "Cisteni/Cleaner s"),
    8: (0, "-", "Navrat do skladu/ToStorageGlass"),
    9: (6, "-", "Drink k zakaznikovi/ToCustomer"),
    10: (7, "ProcessCellID", "Transport (ParameterProcess1=ProcessCellID)"),
}

# --- TRecipeInfo / TRecipeStep fields from NFC_reader.h ---
TAG_INFO_FIELDS = [
    ("TRecipeInfo", "ID", "uint8_t", "Recipe identifier", "NFC_reader.h:21"),
    ("TRecipeInfo", "NumOfDrinks", "uint16_t", "Number of drinks produced", "NFC_reader.h:22"),
    ("TRecipeInfo", "RecipeSteps", "uint8_t", "Total number of recipe steps", "NFC_reader.h:23"),
    ("TRecipeInfo", "ActualRecipeStep", "uint8_t", "Current recipe step index", "NFC_reader.h:24"),
    ("TRecipeInfo", "ActualBudget", "uint16_t", "Current budget remaining", "NFC_reader.h:25"),
    ("TRecipeInfo", "Parameters", "uint8_t", "Recipe parameters", "NFC_reader.h:26"),
    ("TRecipeInfo", "RightNumber", "uint8_t", "Validation number (255-ID)", "NFC_reader.h:27"),
    ("TRecipeInfo", "RecipeDone", "bool", "Recipe completion flag", "NFC_reader.h:28"),
    ("TRecipeInfo", "CheckSum", "uint16_t", "Checksum (must be last)", "NFC_reader.h:29"),
    ("TRecipeStep", "ID", "uint8_t", "Step identifier", "NFC_reader.h:33"),
    ("TRecipeStep", "NextID", "uint8_t", "Next step identifier", "NFC_reader.h:34"),
    ("TRecipeStep", "TypeOfProcess", "uint8_t", "ProcessTypes enum", "NFC_reader.h:35"),
    ("TRecipeStep", "ParameterProcess1", "uint8_t", "AlcoholType/NonAlcoholType or duration or ProcessCellID", "NFC_reader.h:36"),
    ("TRecipeStep", "ParameterProcess2", "uint16_t", "Volume ml or duration", "NFC_reader.h:37"),
    ("TRecipeStep", "PriceForTransport", "uint8_t", "Transport cost", "NFC_reader.h:38"),
    ("TRecipeStep", "TransportCellID", "uint8_t", "Transport cell identifier", "NFC_reader.h:39"),
    ("TRecipeStep", "TransportCellReservationID", "uint16_t", "Transport reservation ID", "NFC_reader.h:40"),
    ("TRecipeStep", "PriceForProcess", "uint8_t", "Process cost", "NFC_reader.h:41"),
    ("TRecipeStep", "ProcessCellID", "uint8_t", "Process cell identifier", "NFC_reader.h:42"),
    ("TRecipeStep", "ProcessCellReservationID", "uint16_t", "Process reservation ID", "NFC_reader.h:43"),
    ("TRecipeStep", "TimeOfProcess", "UA_DateTime", "Process execution time", "NFC_reader.h:44"),
    ("TRecipeStep", "TimeOfTransport", "UA_DateTime", "Transport execution time", "NFC_reader.h:45"),
    ("TRecipeStep", "NeedForTransport", "bool", "Transport required flag", "NFC_reader.h:46"),
    ("TRecipeStep", "IsTransport", "bool", "Transport completed flag", "NFC_reader.h:47"),
    ("TRecipeStep", "IsProcess", "bool", "Process completed flag", "NFC_reader.h:48"),
    ("TRecipeStep", "IsStepDone", "bool", "Step completed flag", "NFC_reader.h:49"),
]


def type_of_process_for_step_type(step_type: int, param: int) -> tuple:
    """Return (TypeOfProcess, ParameterProcess1, ParameterProcess2) for GetRecipeStepByNumber(step_type, param)."""
    info = STEP_TYPE_INFO.get(step_type)
    if not info:
        return (0, 0, 0)
    type_val, p1_meaning, _ = info
    if step_type in (1, 2, 3):
        p1 = step_type - 1  # Vodka=0, Rum=1, Goralka=2
        return (type_val, p1, param)
    if step_type in (4, 5):
        p1 = step_type - 4  # Water=0, Cola=1
        return (type_val, p1, param)
    if step_type == 6:
        return (type_val, param, 0)  # Shaker: ParameterProcess1 = duration
    if step_type == 7:
        return (type_val, param, 0)  # Cleaner: ParameterProcess1 = duration
    if step_type in (8, 9):
        return (type_val, 0, 0)
    if step_type == 10:
        return (type_val, param, 0)  # Transport: ParameterProcess1 = ProcessCellID
    return (0, 0, 0)


def type_of_process_name(val: int) -> str:
    for v, name in PROCESS_TYPES:
        if v == val:
            return name
    return ""


def build_steps_rows():
    rows = []
    for r in RECIPES:
        for i, (step_type, param) in enumerate(r["Steps"]):
            top, p1, p2 = type_of_process_for_step_type(step_type, param)
            step_label = STEP_TYPE_INFO.get(step_type, (None, "", ""))[2] if step_type else ""
            rows.append({
                "RecipeNumber": r["RecipeNumber"],
                "StepIndex": i,
                "StepName/Label": step_label,
                "TypeOfProcess": top,
                "TypeOfProcessName": type_of_process_name(top),
                "ParameterProcess1": p1,
                "ParameterProcess2": p2,
                "ProcessCellID": 0,
                "TransportCellID": 0,
                "SourceLocation": f"NFC_recipes.c GetRecipeStepByNumber({step_type},{param}) -> step in GetCardInfoByNumber",
            })
    return rows


def main():
    out_dir = Path(__file__).resolve().parent.parent
    xlsx_path = out_dir / "firmware_recipes.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Recipes
    ws1 = wb.create_sheet("Recipes", 0)
    ws1.append(["RecipeNumber", "RecipeName", "NumSteps", "Description/Notes", "SourceLocation"])
    for r in RECIPES:
        notes = r["Description"]
        if r.get("ActualBudget") is not None:
            notes += f" ActualBudget={r['ActualBudget']}."
        ws1.append([r["RecipeNumber"], r["RecipeName"], r["NumSteps"], notes, r["SourceLocation"]])
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    # Sheet 2: Steps
    ws2 = wb.create_sheet("Steps", 1)
    step_rows = build_steps_rows()
    headers = ["RecipeNumber", "StepIndex", "StepName/Label", "TypeOfProcess", "TypeOfProcessName",
               "ParameterProcess1", "ParameterProcess2", "ProcessCellID", "TransportCellID", "SourceLocation"]
    ws2.append(headers)
    for row in step_rows:
        ws2.append([row[h] for h in headers])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    # Sheet 3: Enums
    ws3 = wb.create_sheet("Enums", 2)
    ws3.append(["EnumName", "Value", "Name", "SourceLocation"])
    for v, name in PROCESS_TYPES:
        ws3.append(["ProcessTypes", v, name, "NFC_recipes.h:28-38"])
    for v, name in ALCOHOL_TYPE:
        ws3.append(["AlcoholType", v, name, "NFC_recipes.h:40-44"])
    for v, name in NON_ALCOHOL_TYPE:
        ws3.append(["NonAlcoholType", v, name, "NFC_recipes.h:46-49"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    # Sheet 4: TagInfoFields
    ws4 = wb.create_sheet("TagInfoFields", 3)
    ws4.append(["Struct", "FieldName", "Type/Size", "Meaning", "SourceLocation"])
    for struct, field, typ, meaning, loc in TAG_INFO_FIELDS:
        ws4.append([struct, field, typ, meaning, loc])
    for cell in ws4[1]:
        cell.font = Font(bold=True)

    wb.save(xlsx_path)
    print(f"Written {xlsx_path}")
    return xlsx_path


if __name__ == "__main__":
    main()
